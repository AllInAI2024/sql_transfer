from typing import List
from io import BytesIO

from fastapi import APIRouter, Request, UploadFile, File, HTTPException, Depends
from fastapi.responses import HTMLResponse, JSONResponse
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import get_settings, templates
from app.database import get_db
from app.models import IntermediateScript

settings = get_settings()

router_intermediate = APIRouter()


@router_intermediate.get("/", response_class=HTMLResponse)
async def intermediate_scripts(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="intermediate.html",
        context={
            "title": "中间表",
            "app_name": settings.APP_NAME,
            "app_version": settings.APP_VERSION,
            "active_menu": "intermediate"
        }
    )


@router_intermediate.get("/list")
async def list_intermediate_scripts(db: Session = Depends(get_db)):
    scripts = db.query(IntermediateScript).order_by(IntermediateScript.created_at.desc()).all()
    
    result = []
    for script in scripts:
        result.append({
            "id": script.id,
            "intermediate_table_name": script.intermediate_table_name,
            "script": script.script,
            "description": script.description,
            "created_at": script.created_at.isoformat() if script.created_at else None,
            "updated_at": script.updated_at.isoformat() if script.updated_at else None
        })
    
    return JSONResponse(content={"success": True, "data": result})


@router_intermediate.get("/{script_id}")
async def get_intermediate_script(script_id: int, db: Session = Depends(get_db)):
    script = db.query(IntermediateScript).filter(IntermediateScript.id == script_id).first()
    
    if not script:
        raise HTTPException(status_code=404, detail="中间表脚本不存在")
    
    result = {
        "id": script.id,
        "intermediate_table_name": script.intermediate_table_name,
        "script": script.script,
        "description": script.description,
        "created_at": script.created_at.isoformat() if script.created_at else None,
        "updated_at": script.updated_at.isoformat() if script.updated_at else None
    }
    
    return JSONResponse(content={"success": True, "data": result})


@router_intermediate.post("/import")
async def import_intermediate_scripts(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件 (.xlsx 或 .xls)")
    
    try:
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents), data_only=True)
        sheet = workbook.active
        
        if sheet.max_row < 2:
            raise HTTPException(status_code=400, detail="Excel 文件中没有数据")
        
        imported_count = 0
        updated_count = 0
        errors = []
        
        for row_idx in range(2, sheet.max_row + 1):
            table_name = sheet.cell(row=row_idx, column=1).value
            script_content = sheet.cell(row=row_idx, column=2).value
            
            if not table_name or not script_content:
                errors.append(f"第 {row_idx} 行：中间表名或脚本为空")
                continue
            
            table_name = str(table_name).strip()
            script_content = str(script_content).strip()
            
            if not table_name or not script_content:
                errors.append(f"第 {row_idx} 行：中间表名或脚本为空")
                continue
            
            existing_script = db.query(IntermediateScript).filter(
                IntermediateScript.intermediate_table_name == table_name
            ).first()
            
            if existing_script:
                existing_script.script = script_content
                updated_count += 1
            else:
                new_script = IntermediateScript(
                    intermediate_table_name=table_name,
                    script=script_content
                )
                db.add(new_script)
                imported_count += 1
        
        db.commit()
        
        message = f"导入完成：新增 {imported_count} 条，更新 {updated_count} 条"
        if errors:
            message += f"，{len(errors)} 条数据有错误"
        
        return JSONResponse(content={
            "success": True,
            "message": message,
            "imported_count": imported_count,
            "updated_count": updated_count,
            "errors": errors
        })
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")


@router_intermediate.delete("/{script_id}")
async def delete_intermediate_script(script_id: int, db: Session = Depends(get_db)):
    script = db.query(IntermediateScript).filter(IntermediateScript.id == script_id).first()
    
    if not script:
        raise HTTPException(status_code=404, detail="中间表脚本不存在")
    
    db.delete(script)
    db.commit()
    
    return JSONResponse(content={"success": True, "message": "删除成功"})
