from typing import Dict, List, Optional
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.config import get_settings, templates
from app.database import get_db
from app.models.models import Dialect, IntermediateScript, Task, VisualizationScript
from app.services.anonymization_service import anonymize_sql, deanonymize_sql
from app.services.visualization_import_service import process_visualization_script_import

settings = get_settings()

router_anonymization = APIRouter()


class AnonymizeRequest(BaseModel):
    sql: str
    exclude_list: Optional[str] = None
    dialect: Optional[str] = None
    remove_comments: Optional[bool] = True


class DeanonymizeRequest(BaseModel):
    sql: str
    mapping: Dict[str, Dict[str, str]]
    dialect: Optional[str] = None


class AnonymizeResponse(BaseModel):
    success: bool
    anonymized_sql: Optional[str] = None
    mapping: Optional[Dict[str, Dict[str, str]]] = None
    error: Optional[str] = None


class DeanonymizeResponse(BaseModel):
    success: bool
    deanonymized_sql: Optional[str] = None
    error: Optional[str] = None


class ConfigResponse(BaseModel):
    success: bool
    exclude_list: str
    dialect: str


class TaskCreateRequest(BaseModel):
    name: str
    description: Optional[str] = None
    intermediate_dialect_id: int
    visualization_dialect_id: int
    converted_dialect_id: int


class TaskResponse(BaseModel):
    id: int
    name: str
    description: Optional[str] = None
    intermediate_dialect_id: Optional[int] = None
    intermediate_dialect_name: Optional[str] = None
    visualization_dialect_id: Optional[int] = None
    visualization_dialect_name: Optional[str] = None
    converted_dialect_id: Optional[int] = None
    converted_dialect_name: Optional[str] = None
    created_at: Optional[str] = None
    updated_at: Optional[str] = None
    visualization_script_count: int = 0


class TaskListResponse(BaseModel):
    success: bool
    tasks: List[TaskResponse] = []
    error: Optional[str] = None


class TaskDeleteResponse(BaseModel):
    success: bool
    error: Optional[str] = None


class DialectOption(BaseModel):
    id: int
    name: str
    display_name: str


class DialectListResponse(BaseModel):
    success: bool
    dialects: List[DialectOption] = []


@router_anonymization.get("/", response_class=HTMLResponse)
async def anonymization_page(request: Request, db: Session = Depends(get_db)):
    dialects = db.query(Dialect).filter(Dialect.is_enabled == True).order_by(Dialect.sort_order).all()
    
    tasks = db.query(Task).order_by(Task.updated_at.desc()).all()
    
    task_responses = []
    for task in tasks:
        vis_count = len(task.visualization_scripts) if task.visualization_scripts else 0
        task_responses.append(TaskResponse(
            id=task.id,
            name=task.name,
            description=task.description,
            intermediate_dialect_id=task.intermediate_dialect_id,
            intermediate_dialect_name=task.intermediate_dialect.display_name if task.intermediate_dialect else None,
            visualization_dialect_id=task.visualization_dialect_id,
            visualization_dialect_name=task.visualization_dialect.display_name if task.visualization_dialect else None,
            converted_dialect_id=task.converted_dialect_id,
            converted_dialect_name=task.converted_dialect.display_name if task.converted_dialect else None,
            created_at=task.created_at.strftime('%Y-%m-%d %H:%M:%S') if task.created_at else None,
            updated_at=task.updated_at.strftime('%Y-%m-%d %H:%M:%S') if task.updated_at else None,
            visualization_script_count=vis_count
        ))
    
    dialect_options = [
        DialectOption(id=d.id, name=d.name, display_name=d.display_name)
        for d in dialects
    ]
    
    return templates.TemplateResponse(
        request=request,
        name="anonymization.html",
        context={
            "title": "匿名化",
            "app_name": settings.APP_NAME,
            "app_version": settings.APP_VERSION,
            "active_menu": "anonymization",
            "dialects": dialect_options,
            "tasks": task_responses,
        }
    )


@router_anonymization.get("/test", response_class=HTMLResponse)
async def anonymization_test_page(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="anonymization_test.html",
        context={
            "title": "匿名化测试",
            "app_name": settings.APP_NAME,
            "app_version": settings.APP_VERSION,
            "active_menu": "anonymization",
            "default_exclude_list": settings.ANONYMIZATION_EXCLUDE_LIST,
        }
    )


@router_anonymization.post("/process", response_model=AnonymizeResponse)
async def process_anonymization(request: AnonymizeRequest):
    try:
        if not request.sql or not request.sql.strip():
            return AnonymizeResponse(
                success=False,
                error="SQL 脚本不能为空"
            )
        
        exclude_list_str = request.exclude_list or settings.ANONYMIZATION_EXCLUDE_LIST
        exclude_list = [item.strip() for item in exclude_list_str.split(",") if item.strip()]
        
        dialect = request.dialect or settings.ANONYMIZATION_DEFAULT_DIALECT
        
        anonymized_sql, mapping = anonymize_sql(
            sql=request.sql,
            exclude_list=exclude_list,
            dialect=dialect,
            remove_comments=request.remove_comments or False
        )
        
        return AnonymizeResponse(
            success=True,
            anonymized_sql=anonymized_sql,
            mapping=mapping
        )
    except Exception as e:
        return AnonymizeResponse(
            success=False,
            error=f"匿名化处理失败: {str(e)}"
        )


@router_anonymization.post("/test/process", response_model=AnonymizeResponse)
async def test_process_anonymization(request: AnonymizeRequest):
    return await process_anonymization(request)


@router_anonymization.post("/test/deanonymize", response_model=DeanonymizeResponse)
async def test_process_deanonymization(request: DeanonymizeRequest):
    try:
        if not request.sql or not request.sql.strip():
            return DeanonymizeResponse(
                success=False,
                error="SQL 脚本不能为空"
            )
        
        if not request.mapping:
            return DeanonymizeResponse(
                success=False,
                error="编码字典不能为空"
            )
        
        dialect = request.dialect or settings.ANONYMIZATION_DEFAULT_DIALECT
        
        deanonymized_sql = deanonymize_sql(
            sql=request.sql,
            mapping=request.mapping,
            dialect=dialect
        )
        
        return DeanonymizeResponse(
            success=True,
            deanonymized_sql=deanonymized_sql
        )
    except Exception as e:
        return DeanonymizeResponse(
            success=False,
            error=f"反向匿名化处理失败: {str(e)}"
        )


@router_anonymization.get("/config", response_model=ConfigResponse)
async def get_anonymization_config():
    return ConfigResponse(
        success=True,
        exclude_list=settings.ANONYMIZATION_EXCLUDE_LIST,
        dialect=settings.ANONYMIZATION_DEFAULT_DIALECT
    )


@router_anonymization.get("/history")
async def get_anonymization_history():
    return {"message": "获取匿名化历史记录", "data": []}


@router_anonymization.get("/api/dialects", response_model=DialectListResponse)
async def get_dialects(db: Session = Depends(get_db)):
    try:
        dialects = db.query(Dialect).filter(Dialect.is_enabled == True).order_by(Dialect.sort_order).all()
        return DialectListResponse(
            success=True,
            dialects=[
                DialectOption(id=d.id, name=d.name, display_name=d.display_name)
                for d in dialects
            ]
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router_anonymization.post("/api/tasks", response_model=TaskResponse)
async def create_task(request: TaskCreateRequest, db: Session = Depends(get_db)):
    try:
        if not request.name or not request.name.strip():
            raise HTTPException(status_code=400, detail="任务名称不能为空")
        
        if request.intermediate_dialect_id != request.converted_dialect_id:
            raise HTTPException(
                status_code=400, 
                detail="中间表方言和转换脚本方言必须相同（因为两者都是基于原始数据的脚本）"
            )
        
        intermediate_dialect = db.query(Dialect).filter(Dialect.id == request.intermediate_dialect_id).first()
        visualization_dialect = db.query(Dialect).filter(Dialect.id == request.visualization_dialect_id).first()
        converted_dialect = db.query(Dialect).filter(Dialect.id == request.converted_dialect_id).first()
        
        if not intermediate_dialect:
            raise HTTPException(status_code=400, detail="中间表方言不存在")
        if not visualization_dialect:
            raise HTTPException(status_code=400, detail="可视化脚本方言不存在")
        if not converted_dialect:
            raise HTTPException(status_code=400, detail="转换脚本方言不存在")
        
        existing_task = db.query(Task).filter(Task.name == request.name.strip()).first()
        if existing_task:
            raise HTTPException(status_code=400, detail=f"任务名称 '{request.name}' 已存在")
        
        task = Task(
            name=request.name.strip(),
            description=request.description,
            intermediate_dialect_id=request.intermediate_dialect_id,
            visualization_dialect_id=request.visualization_dialect_id,
            converted_dialect_id=request.converted_dialect_id
        )
        
        db.add(task)
        db.commit()
        db.refresh(task)
        
        return TaskResponse(
            id=task.id,
            name=task.name,
            description=task.description,
            intermediate_dialect_id=task.intermediate_dialect_id,
            intermediate_dialect_name=intermediate_dialect.display_name,
            visualization_dialect_id=task.visualization_dialect_id,
            visualization_dialect_name=visualization_dialect.display_name,
            converted_dialect_id=task.converted_dialect_id,
            converted_dialect_name=converted_dialect.display_name,
            created_at=task.created_at.strftime('%Y-%m-%d %H:%M:%S') if task.created_at else None,
            updated_at=task.updated_at.strftime('%Y-%m-%d %H:%M:%S') if task.updated_at else None,
            visualization_script_count=0
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router_anonymization.get("/api/tasks", response_model=TaskListResponse)
async def list_tasks(db: Session = Depends(get_db)):
    try:
        tasks = db.query(Task).order_by(Task.updated_at.desc()).all()
        
        task_responses = []
        for task in tasks:
            vis_count = len(task.visualization_scripts) if task.visualization_scripts else 0
            task_responses.append(TaskResponse(
                id=task.id,
                name=task.name,
                description=task.description,
                intermediate_dialect_id=task.intermediate_dialect_id,
                intermediate_dialect_name=task.intermediate_dialect.display_name if task.intermediate_dialect else None,
                visualization_dialect_id=task.visualization_dialect_id,
                visualization_dialect_name=task.visualization_dialect.display_name if task.visualization_dialect else None,
                converted_dialect_id=task.converted_dialect_id,
                converted_dialect_name=task.converted_dialect.display_name if task.converted_dialect else None,
                created_at=task.created_at.strftime('%Y-%m-%d %H:%M:%S') if task.created_at else None,
                updated_at=task.updated_at.strftime('%Y-%m-%d %H:%M:%S') if task.updated_at else None,
                visualization_script_count=vis_count
            ))
        
        return TaskListResponse(success=True, tasks=task_responses)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router_anonymization.delete("/api/tasks/{task_id}", response_model=TaskDeleteResponse)
async def delete_task(task_id: int, db: Session = Depends(get_db)):
    try:
        task = db.query(Task).filter(Task.id == task_id).first()
        
        if not task:
            raise HTTPException(status_code=404, detail=f"任务 ID {task_id} 不存在")
        
        vis_script_count = db.query(VisualizationScript).filter(
            VisualizationScript.task_id == task_id
        ).count()
        
        db.delete(task)
        db.commit()
        
        return TaskDeleteResponse(
            success=True,
            error=f"任务已删除，同时删除了 {vis_script_count} 个关联的可视化脚本"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router_anonymization.get("/task/{task_id}", response_class=HTMLResponse)
async def task_detail_page(task_id: int, request: Request, db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    
    visualization_scripts = db.query(VisualizationScript).filter(
        VisualizationScript.task_id == task_id
    ).order_by(VisualizationScript.updated_at.desc()).all()
    
    script_list = []
    for vs in visualization_scripts:
        script_list.append({
            "id": vs.id,
            "name": vs.name,
            "intermediate_table_names": vs.intermediate_table_names,
            "description": vs.description,
            "has_integrated_script": bool(vs.integrated_script),
            "has_anonymized_script": bool(vs.anonymized_integrated_script),
            "has_converted_script": bool(vs.converted_script),
            "created_at": vs.created_at.strftime('%Y-%m-%d %H:%M:%S') if vs.created_at else None,
            "updated_at": vs.updated_at.strftime('%Y-%m-%d %H:%M:%S') if vs.updated_at else None,
        })
    
    return templates.TemplateResponse(
        request=request,
        name="task_detail.html",
        context={
            "title": f"任务详情 - {task.name}",
            "app_name": settings.APP_NAME,
            "app_version": settings.APP_VERSION,
            "active_menu": "anonymization",
            "task": {
                "id": task.id,
                "name": task.name,
                "description": task.description,
                "intermediate_dialect_name": task.intermediate_dialect.display_name if task.intermediate_dialect else None,
                "visualization_dialect_name": task.visualization_dialect.display_name if task.visualization_dialect else None,
                "converted_dialect_name": task.converted_dialect.display_name if task.converted_dialect else None,
                "created_at": task.created_at.strftime('%Y-%m-%d %H:%M:%S') if task.created_at else None,
                "updated_at": task.updated_at.strftime('%Y-%m-%d %H:%M:%S') if task.updated_at else None,
            },
            "scripts": script_list,
        }
    )


@router_anonymization.get("/api/task/{task_id}/scripts")
async def list_task_scripts(task_id: int, db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    
    scripts = db.query(VisualizationScript).filter(
        VisualizationScript.task_id == task_id
    ).order_by(VisualizationScript.updated_at.desc()).all()
    
    result = []
    for vs in scripts:
        result.append({
            "id": vs.id,
            "name": vs.name,
            "intermediate_table_names": vs.intermediate_table_names,
            "description": vs.description,
            "has_integrated_script": bool(vs.integrated_script),
            "has_anonymized_script": bool(vs.anonymized_integrated_script),
            "has_converted_script": bool(vs.converted_script),
            "created_at": vs.created_at.strftime('%Y-%m-%d %H:%M:%S') if vs.created_at else None,
            "updated_at": vs.updated_at.strftime('%Y-%m-%d %H:%M:%S') if vs.updated_at else None,
        })
    
    return JSONResponse(content={
        "success": True,
        "data": result,
        "task_name": task.name
    })


@router_anonymization.post("/api/task/{task_id}/import")
async def import_visualization_scripts(
    task_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件 (.xlsx 或 .xls)")
    
    try:
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents), data_only=True)
        sheet = workbook.active
        
        if sheet.max_row < 2:
            raise HTTPException(status_code=400, detail="Excel 文件中没有数据")
        
        intermediate_scripts = db.query(IntermediateScript).all()
        intermediate_scripts_dict = {
            iscript.intermediate_table_name.lower(): iscript.script
            for iscript in intermediate_scripts
        }
        
        scripts_to_import = []
        errors = []
        validation_errors = []
        
        for row_idx in range(2, sheet.max_row + 1):
            script_name = sheet.cell(row=row_idx, column=1).value
            visualization_script_content = sheet.cell(row=row_idx, column=2).value
            
            if not script_name or not visualization_script_content:
                if script_name or visualization_script_content:
                    errors.append(f"第 {row_idx} 行：脚本名称或可视化脚本为空")
                continue
            
            script_name = str(script_name).strip()
            visualization_script_content = str(visualization_script_content).strip()
            
            if not script_name or not visualization_script_content:
                errors.append(f"第 {row_idx} 行：脚本名称或可视化脚本为空")
                continue
            
            process_result = process_visualization_script_import(
                script_name=script_name,
                visualization_script=visualization_script_content,
                existing_intermediate_scripts=intermediate_scripts_dict
            )
            
            if process_result['missing_tables']:
                validation_errors.append(
                    f"脚本 '{script_name}' 中引用的中间表未找到: {', '.join(process_result['missing_tables'])}"
                )
            
            scripts_to_import.append({
                'row_idx': row_idx,
                'name': script_name,
                'visualization_script': visualization_script_content,
                'process_result': process_result
            })
        
        if validation_errors:
            raise HTTPException(
                status_code=400,
                detail=f"校验失败：{validation_errors[0]}"
            )
        
        imported_count = 0
        updated_count = 0
        warning_messages = []
        
        for script_data in scripts_to_import:
            script_name = script_data['name']
            visualization_script_content = script_data['visualization_script']
            process_result = script_data['process_result']
            
            existing_script = db.query(VisualizationScript).filter(
                VisualizationScript.task_id == task_id,
                VisualizationScript.name == script_name
            ).first()
            
            if existing_script:
                existing_script.visualization_script = visualization_script_content
                existing_script.intermediate_table_names = process_result['intermediate_table_names']
                existing_script.integrated_script = process_result['integrated_script']
                updated_count += 1
            else:
                new_script = VisualizationScript(
                    task_id=task_id,
                    name=script_name,
                    visualization_script=visualization_script_content,
                    intermediate_table_names=process_result['intermediate_table_names'],
                    integrated_script=process_result['integrated_script']
                )
                db.add(new_script)
                imported_count += 1
        
        db.commit()
        
        message = f"导入完成：新增 {imported_count} 条，更新 {updated_count} 条"
        if warning_messages:
            message += f"，{len(warning_messages)} 个警告"
        
        return JSONResponse(content={
            "success": True,
            "message": message,
            "imported_count": imported_count,
            "updated_count": updated_count,
            "warnings": warning_messages,
            "errors": errors
        })
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")


@router_anonymization.delete("/api/script/{script_id}")
async def delete_visualization_script(script_id: int, db: Session = Depends(get_db)):
    script = db.query(VisualizationScript).filter(VisualizationScript.id == script_id).first()
    
    if not script:
        raise HTTPException(status_code=404, detail="可视化脚本不存在")
    
    db.delete(script)
    db.commit()
    
    return JSONResponse(content={
        "success": True,
        "message": "删除成功"
    })


@router_anonymization.get("/api/script/{script_id}")
async def get_visualization_script_detail(script_id: int, db: Session = Depends(get_db)):
    script = db.query(VisualizationScript).filter(VisualizationScript.id == script_id).first()
    
    if not script:
        raise HTTPException(status_code=404, detail="可视化脚本不存在")
    
    return JSONResponse(content={
        "success": True,
        "data": {
            "id": script.id,
            "name": script.name,
            "description": script.description,
            "intermediate_table_names": script.intermediate_table_names,
            "visualization_script": script.visualization_script,
            "integrated_script": script.integrated_script,
            "anonymized_integrated_script": script.anonymized_integrated_script,
            "anonymization_mapping": script.anonymization_mapping,
            "converted_script": script.converted_script,
            "created_at": script.created_at.strftime('%Y-%m-%d %H:%M:%S') if script.created_at else None,
            "updated_at": script.updated_at.strftime('%Y-%m-%d %H:%M:%S') if script.updated_at else None,
        }
    })


def _perform_anonymization(script: VisualizationScript, db: Session) -> Dict:
    """
    执行单个脚本的匿名化逻辑（可复用的核心方法）
    
    Args:
        script: VisualizationScript 实例
        db: 数据库会话
        
    Returns:
        Dict: 包含 success, message, data 或 error 信息
    """
    if not script.integrated_script:
        return {
            "success": False,
            "script_id": script.id,
            "script_name": script.name,
            "error": "整合脚本为空，无法进行匿名化"
        }
    
    try:
        exclude_list_str = settings.ANONYMIZATION_EXCLUDE_LIST
        exclude_list = [item.strip() for item in exclude_list_str.split(",") if item.strip()]
        
        dialect = settings.ANONYMIZATION_DEFAULT_DIALECT
        
        anonymized_sql, mapping = anonymize_sql(
            sql=script.integrated_script,
            exclude_list=exclude_list,
            dialect=dialect,
            remove_comments=True
        )
        
        script.anonymized_integrated_script = anonymized_sql
        script.anonymization_mapping = str(mapping)
        
        return {
            "success": True,
            "script_id": script.id,
            "script_name": script.name,
            "anonymized_script": anonymized_sql,
            "mapping": mapping
        }
        
    except Exception as e:
        return {
            "success": False,
            "script_id": script.id,
            "script_name": script.name,
            "error": str(e)
        }


@router_anonymization.post("/api/script/{script_id}/anonymize")
async def anonymize_visualization_script(script_id: int, db: Session = Depends(get_db)):
    script = db.query(VisualizationScript).filter(VisualizationScript.id == script_id).first()
    
    if not script:
        raise HTTPException(status_code=404, detail="可视化脚本不存在")
    
    result = _perform_anonymization(script, db)
    
    if result["success"]:
        db.commit()
        return JSONResponse(content={
            "success": True,
            "message": "匿名化完成",
            "data": {
                "anonymized_script": result["anonymized_script"],
                "mapping": result["mapping"]
            }
        })
    else:
        db.rollback()
        raise HTTPException(status_code=400, detail=result["error"])


@router_anonymization.post("/api/task/{task_id}/anonymize-all")
async def anonymize_all_visualization_scripts(task_id: int, db: Session = Depends(get_db)):
    """
    批量匿名化指定任务下的所有可视化脚本
    
    逻辑：
    1. 获取该任务下的所有可视化脚本
    2. 过滤掉已经匿名化的脚本（anonymized_integrated_script 不为空）
    3. 对每个未匿名化的脚本执行匿名化
    4. 返回统计信息
    """
    task = db.query(Task).filter(Task.id == task_id).first()
    
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    
    scripts = db.query(VisualizationScript).filter(
        VisualizationScript.task_id == task_id
    ).order_by(VisualizationScript.id.asc()).all()
    
    if not scripts:
        return JSONResponse(content={
            "success": True,
            "message": "该任务下没有可视化脚本",
            "data": {
                "total": 0,
                "processed": 0,
                "skipped": 0,
                "failed": 0,
                "results": []
            }
        })
    
    total_count = len(scripts)
    skipped_count = 0
    success_count = 0
    failed_count = 0
    results = []
    
    for script in scripts:
        if script.anonymized_integrated_script:
            skipped_count += 1
            results.append({
                "script_id": script.id,
                "script_name": script.name,
                "status": "skipped",
                "message": "已经匿名化，跳过"
            })
            continue
        
        result = _perform_anonymization(script, db)
        
        if result["success"]:
            success_count += 1
            results.append({
                "script_id": result["script_id"],
                "script_name": result["script_name"],
                "status": "success",
                "message": "匿名化成功"
            })
        else:
            failed_count += 1
            results.append({
                "script_id": result["script_id"],
                "script_name": result["script_name"],
                "status": "failed",
                "message": result["error"]
            })
    
    db.commit()
    
    message = f"批量匿名化完成：共 {total_count} 个，成功 {success_count} 个，跳过 {skipped_count} 个，失败 {failed_count} 个"
    
    return JSONResponse(content={
        "success": True,
        "message": message,
        "data": {
            "total": total_count,
            "processed": success_count,
            "skipped": skipped_count,
            "failed": failed_count,
            "results": results
        }
    })
